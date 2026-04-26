#!/usr/bin/env python3
"""
export_excel.py

Builds a single consolidated XLSX file containing the full investment history
in human-readable form, so that the data can be re-imported into other apps.

Tabs:
    Resumo         — high-level KPIs (V_end atual, total investido, TWR, ...)
    TWR_Mensal     — monthly TWR, V_end, cashflow, income (Total Return)
    Vs_Benchmarks  — TWR cumulative vs IBOV, IVVB11, S&P500, CDI
    Holdings_Atual — final positions (ticker, qty, cost_basis)
    Holdings_Mes   — ticker × month_end qty matrix
    Trades         — every BUY/SELL trade
    Proventos      — every dividend/JCP/rendimento/resgate
    Splits         — corporate events (agrupamentos/desdobramentos)
    Aliases        — ticker renames (TRPL4->ISAE4 etc.)
    Precos         — monthly close cache (ticker × month_end)

Output: investimentos_consolidado.xlsx
"""
from __future__ import annotations

import argparse
import csv
import sys
from collections import defaultdict
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing openpyxl. pip install openpyxl")
    sys.exit(1)


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
TITLE_FONT = Font(bold=True, size=14, color="305496")
THIN = Side(border_style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM_FMTS = {
    "money_brl": 'R$ #,##0.00;[Red]-R$ #,##0.00',
    "money_int": 'R$ #,##0;[Red]-R$ #,##0',
    "pct":       '0.00%;[Red]-0.00%',
    "qty":       '#,##0.######',
    "int":       '#,##0',
}


def _read_csv(path: str) -> list[dict]:
    try:
        with open(path, encoding="utf-8") as f:
            return list(csv.DictReader(f))
    except FileNotFoundError:
        return []


def _autosize(ws, max_width: int = 40):
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        longest = 0
        for c in col_cells:
            v = "" if c.value is None else str(c.value)
            longest = max(longest, len(v))
        ws.column_dimensions[col_letter].width = min(longest + 2, max_width)


def _write_table(ws, headers: list[str], rows: list[list],
                 number_formats: dict[str, str] | None = None,
                 start_row: int = 1):
    """Writes a table with styled header. Returns row after last data row."""
    number_formats = number_formats or {}
    # Header
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    # Data
    for i, row in enumerate(rows, start=start_row + 1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.border = BORDER
            fmt = number_formats.get(headers[j - 1])
            if fmt:
                c.number_format = fmt
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return start_row + 1 + len(rows)


def _f(s, default=0.0):
    try:
        return float(s) if s not in (None, "") else default
    except (TypeError, ValueError):
        return default


def sheet_resumo(wb, twr, full, holdings_final, proventos, trades,
                 rf_final, us_final):
    ws = wb.create_sheet("Resumo", 0)
    ws["A1"] = "Resumo da Carteira"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    if not twr:
        return

    last = twr[-1]
    first = twr[0]
    n_months = len(twr)
    years = n_months / 12

    v_end_rv = _f(last.get("v_end_rv", last.get("v_end", 0)))
    v_end_rf_liq = _f(last.get("v_end_rf_liquido", 0))
    v_end_us_liq = _f(last.get("v_end_us_liquido", 0))
    v_end_total_liq = _f(last.get("v_end_liquido", last.get("v_end", 0)))
    v_end_total_brt = _f(last.get("v_end_bruto", v_end_total_liq))
    twr_cum_liq = _f(last["return_cum"])
    twr_cum_brt = _f(last.get("return_cum_bruto", twr_cum_liq))
    twr_ann_liq = (1 + twr_cum_liq) ** (1 / years) - 1 if years > 0 else 0
    twr_ann_brt = (1 + twr_cum_brt) ** (1 / years) - 1 if years > 0 else 0
    total_invested = sum(_f(r["cashflow"]) for r in twr)
    total_income = sum(_f(r["income"]) for r in twr)
    total_ir_rf = sum(_f(r.get("rf_ir_estimado", 0)) for r in twr)

    cost_basis_total = sum(_f(h["cost_basis"]) for h in holdings_final
                           if h["ticker"] != "__EMPTY__")
    pnl_unreal = v_end_rv - cost_basis_total
    realized = sum(_f(h.get("realized_pnl", 0)) for h in holdings_final
                   if h["ticker"] != "__EMPTY__")

    n_buys = sum(1 for t in trades if t.get("operacao") == "BUY")
    n_sells = sum(1 for t in trades if t.get("operacao") == "SELL")
    n_provs = len(proventos)
    n_tickers_active = sum(1 for h in holdings_final
                           if h["ticker"] != "__EMPTY__"
                           and abs(_f(h["qty"])) > 1e-6)
    n_rf_active = sum(1 for r in rf_final if abs(_f(r.get("qty", 0))) > 1e-6)
    n_us_active = sum(1 for r in us_final if abs(_f(r.get("qty", 0))) > 1e-6)

    # Benchmarks
    bench_cum = {}
    if full:
        f_last = full[-1]
        for k in ("ibov_cum", "ivvb11_cum", "sp500_cum", "cdi_cum"):
            bench_cum[k] = _f(f_last.get(k, 0))

    rows = [
        ("Período", f"{first['month']}  →  {last['month']}"),
        ("Meses cobertos", n_months),
        ("Anos", round(years, 2)),
        ("", ""),
        ("V_end Renda Variável BR (RV)", v_end_rv),
        ("V_end Renda Fixa (RF, líquido)", v_end_rf_liq),
        ("V_end Ações US (BRL, líquido)", v_end_us_liq),
        ("V_end TOTAL (líquido)",  v_end_total_liq),
        ("V_end TOTAL (bruto)",    v_end_total_brt),
        ("Custo total RV (cost basis)", cost_basis_total),
        ("P&L RV não realizado", pnl_unreal),
        ("P&L RV realizado (vendas)", realized),
        ("", ""),
        ("Aporte líquido total", total_invested),
        ("Proventos recebidos (total)", total_income),
        ("IR estimado retido (RF)", total_ir_rf),
        ("Tickers RV (BR) ativos", n_tickers_active),
        ("Posições RF ativas", n_rf_active),
        ("Tickers US ativos", n_us_active),
        ("Compras (RV)", n_buys),
        ("Vendas (RV)", n_sells),
        ("Lançamentos de proventos", n_provs),
        ("", ""),
        ("TWR cumulativo (líquido)",   twr_cum_liq),
        ("TWR anualizado (líquido)",   twr_ann_liq),
        ("TWR cumulativo (bruto)",     twr_cum_brt),
        ("TWR anualizado (bruto)",     twr_ann_brt),
    ]
    if bench_cum:
        rows.extend([
            ("", ""),
            ("IBOV cumulativo (mesmo período)", bench_cum.get("ibov_cum", 0)),
            ("IVVB11 cumulativo (mesmo período)", bench_cum.get("ivvb11_cum", 0)),
            ("S&P500 cumulativo (mesmo período)", bench_cum.get("sp500_cum", 0)),
            ("CDI cumulativo (mesmo período)", bench_cum.get("cdi_cum", 0)),
        ])

    for i, (label, val) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=i, column=2, value=val)
        if isinstance(val, float):
            if abs(val) < 10 and label.lower().find("twr") >= 0 \
                    or "cumulativo" in label.lower() or "anualizado" in label.lower():
                c.number_format = NUM_FMTS["pct"]
            else:
                c.number_format = NUM_FMTS["money_brl"]
        elif isinstance(val, int):
            c.number_format = NUM_FMTS["int"]

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22

    # Note
    note_row = 3 + len(rows) + 2
    ws.cell(row=note_row, column=1, value="Notas:").font = Font(bold=True)
    notes = [
        "TWR = Time-Weighted Return (Modified Dietz mensal, encadeado).",
        "Total Return inclui dividendos, JCP, rendimentos de FII e resgates.",
        "V_end usa preços de fechamento do último pregão de cada mês.",
        "CDI compostado a partir da série SGS 12 do BCB; IBOV/IVVB/S&P do Yahoo.",
    ]
    for i, n in enumerate(notes):
        ws.cell(row=note_row + 1 + i, column=1,
                value=f"• {n}").alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=note_row + 1 + i, start_column=1,
                       end_row=note_row + 1 + i, end_column=4)


def sheet_twr_mensal(wb, twr):
    ws = wb.create_sheet("TWR_Mensal")
    headers = ["Mês", "Mês fim",
               "V_RV", "V_RF (líq)", "V_TOTAL (líq)", "V_TOTAL (brt)",
               "RV Compras", "RV Vendas", "RV Prêmio opções",
               "RF Aplicação", "RF Resgate (líq)", "RF Resgate (brt)",
               "RF IR estimado",
               "Cashflow (líq)", "Cashflow (brt)",
               "Proventos (líq)", "Proventos (brt)",
               "Dividendos", "JCP", "Rendimentos FII", "Resgates",
               "Retorno mês (líq)", "Retorno mês (brt)",
               "TWR cum (líq)", "TWR cum (brt)",
               "Tickers RV"]
    money_cols = ["V_RV", "V_RF (líq)", "V_TOTAL (líq)", "V_TOTAL (brt)",
                  "RV Compras", "RV Vendas", "RV Prêmio opções",
                  "RF Aplicação", "RF Resgate (líq)", "RF Resgate (brt)",
                  "RF IR estimado",
                  "Cashflow (líq)", "Cashflow (brt)",
                  "Proventos (líq)", "Proventos (brt)",
                  "Dividendos", "JCP", "Rendimentos FII", "Resgates"]
    fmts = {c: NUM_FMTS["money_brl"] for c in money_cols}
    fmts.update({
        "Retorno mês (líq)": NUM_FMTS["pct"],
        "Retorno mês (brt)": NUM_FMTS["pct"],
        "TWR cum (líq)": NUM_FMTS["pct"],
        "TWR cum (brt)": NUM_FMTS["pct"],
        "Tickers RV": NUM_FMTS["int"],
    })
    rows = []
    for r in twr:
        rows.append([
            r["month"], r["month_end"],
            _f(r.get("v_end_rv", r.get("v_end", 0))),
            _f(r.get("v_end_rf_liquido", 0)),
            _f(r.get("v_end_liquido", r.get("v_end", 0))),
            _f(r.get("v_end_bruto", r.get("v_end", 0))),
            _f(r["buy_value"]), _f(r["sell_value"]), _f(r["premium_net"]),
            _f(r.get("rf_buy", 0)), _f(r.get("rf_sell_liquido", 0)),
            _f(r.get("rf_sell_bruto", 0)), _f(r.get("rf_ir_estimado", 0)),
            _f(r["cashflow"]), _f(r.get("cashflow_bruto", r["cashflow"])),
            _f(r.get("income", 0)), _f(r.get("income_bruto", r.get("income", 0))),
            _f(r.get("income_dividendo", 0)), _f(r.get("income_jcp", 0)),
            _f(r.get("income_rendimento", 0)), _f(r.get("income_resgate", 0)),
            _f(r["return_month"]),
            _f(r.get("return_month_bruto", r["return_month"])),
            _f(r["return_cum"]),
            _f(r.get("return_cum_bruto", r["return_cum"])),
            int(_f(r["n_tickers_priced"])),
        ])
    _write_table(ws, headers, rows, fmts)
    _autosize(ws, max_width=18)


def sheet_vs_benchmarks(wb, full):
    ws = wb.create_sheet("Vs_Benchmarks")
    headers = ["Mês", "Mês fim", "V_RV", "V_RF", "V_US", "V_TOTAL",
               "Cashflow", "Proventos",
               "TWR cum (líq)", "TWR cum (brt)",
               "IBOV cumul.", "IVVB11 cumul.",
               "S&P500 cumul.", "CDI cumul."]
    fmts = {
        "V_RV": NUM_FMTS["money_brl"],
        "V_RF": NUM_FMTS["money_brl"],
        "V_US": NUM_FMTS["money_brl"],
        "V_TOTAL": NUM_FMTS["money_brl"],
        "Cashflow": NUM_FMTS["money_brl"],
        "Proventos": NUM_FMTS["money_brl"],
        "TWR cum (líq)": NUM_FMTS["pct"],
        "TWR cum (brt)": NUM_FMTS["pct"],
        "IBOV cumul.": NUM_FMTS["pct"],
        "IVVB11 cumul.": NUM_FMTS["pct"],
        "S&P500 cumul.": NUM_FMTS["pct"],
        "CDI cumul.": NUM_FMTS["pct"],
    }
    rows = []
    for r in full:
        rows.append([
            r["month"], r["month_end"],
            _f(r.get("v_end_rv", 0)), _f(r.get("v_end_rf", 0)),
            _f(r.get("v_end_us", 0)),
            _f(r["v_end"]),
            _f(r["cashflow_month"]), _f(r.get("income_month", 0)),
            _f(r["twr_cum"]), _f(r.get("twr_cum_bruto", r["twr_cum"])),
            _f(r["ibov_cum"]), _f(r["ivvb11_cum"]),
            _f(r["sp500_cum"]), _f(r["cdi_cum"]),
        ])
    _write_table(ws, headers, rows, fmts)
    _autosize(ws, max_width=18)


def sheet_renda_fixa_atual(wb, rf_final):
    """Posição atual em renda fixa (snapshot do último mês)."""
    ws = wb.create_sheet("Renda_Fixa_Atual")
    headers = ["Título", "Código", "Tipo", "Vencimento",
               "Quantidade", "MTM líquido (R$)", "MTM bruto (R$)",
               "Método de avaliação"]
    fmts = {
        "Quantidade": NUM_FMTS["qty"],
        "MTM líquido (R$)": NUM_FMTS["money_brl"],
        "MTM bruto (R$)": NUM_FMTS["money_brl"],
    }
    rows = []
    for r in rf_final:
        rows.append([
            r.get("titulo", ""), r.get("codigo", ""), r.get("tipo", ""),
            r.get("vencimento", ""),
            _f(r.get("qty", 0)),
            _f(r.get("valor_mtm_liquido", 0)),
            _f(r.get("valor_mtm_bruto", 0)),
            r.get("method", ""),
        ])
    rows.sort(key=lambda r: -r[5])  # MTM líquido desc
    _write_table(ws, headers, rows, fmts)
    _autosize(ws, max_width=40)


def sheet_renda_fixa_historico(wb, rf_monthly):
    """Histórico mensal de posições RF (matriz mês × título)."""
    ws = wb.create_sheet("Renda_Fixa_Historico")
    by_me: dict[str, dict[str, float]] = defaultdict(dict)
    titulos: set[str] = set()
    for r in rf_monthly:
        key = r["titulo"]
        by_me[r["month_end"]][key] = _f(r.get("valor_mtm_liquido", 0))
        titulos.add(key)
    months = sorted(by_me)
    titulos_sorted = sorted(titulos)
    headers = ["Mês fim"] + titulos_sorted + ["Total RF (líq)"]
    fmts = {tk: NUM_FMTS["money_brl"] for tk in titulos_sorted}
    fmts["Total RF (líq)"] = NUM_FMTS["money_brl"]
    rows = []
    for me in months:
        row = [me]
        total = 0.0
        for tk in titulos_sorted:
            v = by_me[me].get(tk, 0)
            row.append(v if v else "")
            total += v
        row.append(total)
        rows.append(row)
    _write_table(ws, headers, rows, fmts)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["A"].width = 12


def sheet_us_atual(wb, us_final):
    """Posição atual em ações americanas (snapshot do último mês)."""
    ws = wb.create_sheet("US_Atual")
    headers = ["Ticker", "Quantidade", "Close USD", "Valor USD", "PTAX",
               "Valor BRL"]
    fmts = {
        "Quantidade": NUM_FMTS["qty"],
        "Close USD": '$#,##0.0000',
        "Valor USD": '$#,##0.00',
        "PTAX": '0.0000',
        "Valor BRL": NUM_FMTS["money_brl"],
    }
    rows = []
    for r in us_final:
        rows.append([
            r.get("ticker", ""),
            _f(r.get("qty", 0)),
            _f(r.get("close_usd", 0)),
            _f(r.get("valor_usd", 0)),
            _f(r.get("ptax", 0)),
            _f(r.get("valor_brl", 0)),
        ])
    rows.sort(key=lambda r: -r[5])
    _write_table(ws, headers, rows, fmts)
    _autosize(ws)


def sheet_us_historico(wb, us_summary):
    ws = wb.create_sheet("US_Historico")
    headers = ["Mês fim", "PTAX venda",
               "V_stock USD", "Cash USD", "WH acumulado USD",
               "V_end USD (líq)", "V_end USD (brt)",
               "V_end BRL (líq)", "V_end BRL (brt)",
               "Depósito BRL", "Saque BRL", "Cashflow BRL",
               "Compras USD", "Vendas USD",
               "Dividendo BRL (brt)", "WH BRL", "Income BRL (líq)"]
    fmts = {
        "PTAX venda": '0.0000',
        "V_stock USD": '$#,##0.00',
        "Cash USD": '$#,##0.00',
        "WH acumulado USD": '$#,##0.00',
        "V_end USD (líq)": '$#,##0.00',
        "V_end USD (brt)": '$#,##0.00',
        "V_end BRL (líq)": NUM_FMTS["money_brl"],
        "V_end BRL (brt)": NUM_FMTS["money_brl"],
        "Depósito BRL": NUM_FMTS["money_brl"],
        "Saque BRL": NUM_FMTS["money_brl"],
        "Cashflow BRL": NUM_FMTS["money_brl"],
        "Compras USD": '$#,##0.00',
        "Vendas USD": '$#,##0.00',
        "Dividendo BRL (brt)": NUM_FMTS["money_brl"],
        "WH BRL": NUM_FMTS["money_brl"],
        "Income BRL (líq)": NUM_FMTS["money_brl"],
    }
    rows = []
    for r in us_summary:
        # filter out months where there was no activity
        if (_f(r.get("v_end_usd_liquido", 0)) == 0
                and _f(r.get("cashflow_brl", 0)) == 0
                and _f(r.get("div_gross_brl", 0)) == 0):
            continue
        rows.append([
            r.get("month_end", ""), _f(r.get("ptax_venda", 0)),
            _f(r.get("v_stock_usd", 0)), _f(r.get("cash_usd", 0)),
            _f(r.get("wh_acumulado_usd", 0)),
            _f(r.get("v_end_usd_liquido", 0)), _f(r.get("v_end_usd_bruto", 0)),
            _f(r.get("v_end_brl_liquido", 0)), _f(r.get("v_end_brl_bruto", 0)),
            _f(r.get("deposit_brl", 0)), _f(r.get("withdrawal_brl", 0)),
            _f(r.get("cashflow_brl", 0)),
            _f(r.get("buy_usd", 0)), _f(r.get("sell_usd", 0)),
            _f(r.get("div_gross_brl", 0)), _f(r.get("wh_mes_brl", 0)),
            _f(r.get("income_brl_liquido", 0)),
        ])
    _write_table(ws, headers, rows, fmts)
    _autosize(ws, max_width=18)


def sheet_us_trades(wb, us_trades):
    ws = wb.create_sheet("US_Trades")
    headers = ["Data", "Ticker", "Operação", "Quantidade", "Preço USD",
               "Proceeds USD", "Comissão USD", "Cost Basis USD", "Realized P/L USD"]
    fmts = {
        "Quantidade": NUM_FMTS["qty"],
        "Preço USD": '$#,##0.0000',
        "Proceeds USD": '$#,##0.00',
        "Comissão USD": '$#,##0.0000',
        "Cost Basis USD": '$#,##0.00',
        "Realized P/L USD": '$#,##0.00',
    }
    rows = []
    for t in us_trades:
        rows.append([
            t.get("data", ""), t.get("ticker", ""), t.get("operacao", ""),
            _f(t.get("qty", 0)), _f(t.get("preco_usd", 0)),
            _f(t.get("proceeds_usd", 0)), _f(t.get("commission_usd", 0)),
            _f(t.get("basis_usd", 0)), _f(t.get("realized_pnl_usd", 0)),
        ])
    rows.sort(key=lambda r: (r[0], r[1]))
    _write_table(ws, headers, rows, fmts)
    _autosize(ws)


def sheet_us_dividends(wb, us_divs, us_wh):
    ws = wb.create_sheet("US_Dividends")
    headers = ["Data", "Ticker", "Tipo", "Bruto USD", "WH USD",
               "Líquido USD", "Descrição"]
    fmts = {
        "Bruto USD": '$#,##0.00',
        "WH USD": '$#,##0.00',
        "Líquido USD": '$#,##0.00',
    }
    # Match WH to dividend by (date, ticker)
    wh_map = {}
    for w in us_wh:
        k = (w["data"], w["ticker"])
        wh_map[k] = wh_map.get(k, 0) + _f(w["valor_usd"])
    rows = []
    for d in us_divs:
        gross = _f(d["valor_usd"])
        wh = wh_map.get((d["data"], d["ticker"]), 0)
        net = gross + wh
        rows.append([
            d["data"], d["ticker"], d.get("tipo", ""),
            gross, wh, net, d.get("descricao", ""),
        ])
    rows.sort(key=lambda r: (r[0], r[1]))
    _write_table(ws, headers, rows, fmts)
    _autosize(ws, max_width=50)


def sheet_rf_trades(wb, rf_trades):
    ws = wb.create_sheet("RF_Trades")
    headers = ["Data", "Título", "Código", "Tipo", "Vencimento",
               "Operação", "Quantidade", "PU (R$)", "Valor (R$)", "Instituição"]
    fmts = {
        "Quantidade": NUM_FMTS["qty"],
        "PU (R$)": NUM_FMTS["money_brl"],
        "Valor (R$)": NUM_FMTS["money_brl"],
    }
    rows = []
    for t in rf_trades:
        rows.append([
            t.get("data", ""), t.get("titulo", ""), t.get("codigo", ""),
            t.get("tipo", ""), t.get("vencimento", ""),
            t.get("operacao", ""),
            _f(t.get("qty", 0)), _f(t.get("pu", 0)), _f(t.get("valor", 0)),
            t.get("instituicao", ""),
        ])
    rows.sort(key=lambda r: (r[0], r[1]))
    _write_table(ws, headers, rows, fmts)
    _autosize(ws, max_width=40)


def sheet_holdings_final(wb, holdings):
    ws = wb.create_sheet("Holdings_Atual")
    headers = ["Ticker", "Quantidade", "Custo total (R$)",
               "Preço médio (R$)", "Compras totais (R$)", "P&L realizado (R$)"]
    fmts = {
        "Quantidade": NUM_FMTS["qty"],
        "Custo total (R$)": NUM_FMTS["money_brl"],
        "Preço médio (R$)": NUM_FMTS["money_brl"],
        "Compras totais (R$)": NUM_FMTS["money_brl"],
        "P&L realizado (R$)": NUM_FMTS["money_brl"],
    }
    rows = []
    for h in holdings:
        if h["ticker"] == "__EMPTY__":
            continue
        qty = _f(h["qty"])
        cb = _f(h["cost_basis"])
        avg = cb / qty if qty > 1e-9 else 0
        rows.append([
            h["ticker"], qty, cb, avg,
            _f(h.get("buy_value_total", 0)),
            _f(h.get("realized_pnl", 0)),
        ])
    rows.sort(key=lambda r: -r[2])  # by cost basis desc
    _write_table(ws, headers, rows, fmts)
    _autosize(ws)


def sheet_holdings_mes(wb, holdings_monthly):
    """Wide matrix: rows = month_end, cols = ticker, values = qty."""
    ws = wb.create_sheet("Holdings_Mes")
    by_me: dict[str, dict[str, float]] = defaultdict(dict)
    tickers: set[str] = set()
    for r in holdings_monthly:
        tk = r["ticker"]
        if tk == "__EMPTY__":
            continue
        by_me[r["date"]][tk] = _f(r["qty"])
        tickers.add(tk)
    months = sorted(by_me)
    tickers_sorted = sorted(tickers)
    headers = ["Mês fim"] + tickers_sorted
    fmts = {tk: NUM_FMTS["qty"] for tk in tickers_sorted}
    rows = []
    for me in months:
        row = [me] + [by_me[me].get(tk, 0) for tk in tickers_sorted]
        rows.append(row)
    _write_table(ws, headers, rows, fmts)
    # Don't autosize — too many columns. Set a sane default.
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.column_dimensions["A"].width = 12


def sheet_trades(wb, trades):
    ws = wb.create_sheet("Trades")
    headers = ["Data", "Ticker", "Operação", "Quantidade", "Preço (R$)",
               "Valor (R$)", "Mercado", "Categoria", "Instituição", "Nota"]
    fmts = {
        "Quantidade": NUM_FMTS["qty"],
        "Preço (R$)": NUM_FMTS["money_brl"],
        "Valor (R$)": NUM_FMTS["money_brl"],
    }
    rows = []
    for t in trades:
        rows.append([
            t.get("data", ""), t.get("ticker", ""), t.get("operacao", ""),
            _f(t.get("quantidade", 0)), _f(t.get("preco", 0)),
            _f(t.get("valor", 0)),
            t.get("mercado_origem", ""), t.get("categoria", ""),
            t.get("instituicao", ""), t.get("nota", ""),
        ])
    rows.sort(key=lambda r: (r[0], r[1]))
    _write_table(ws, headers, rows, fmts)
    _autosize(ws, max_width=28)


def sheet_proventos(wb, proventos):
    ws = wb.create_sheet("Proventos")
    headers = ["Data", "Ticker", "Tipo", "Valor (R$)", "Instituição", "Descrição"]
    fmts = {"Valor (R$)": NUM_FMTS["money_brl"]}
    rows = []
    for p in proventos:
        rows.append([
            p["data"], p["ticker"], p["tipo"], _f(p["valor"]),
            p.get("instituicao", ""), p.get("descricao", ""),
        ])
    rows.sort(key=lambda r: (r[0], r[1]))
    _write_table(ws, headers, rows, fmts)
    _autosize(ws, max_width=32)


def sheet_splits(wb, splits):
    ws = wb.create_sheet("Splits")
    headers = ["Data", "Ticker", "Fator", "Razão", "Tipo"]
    fmts = {"Fator": '0.000000'}
    rows = []
    for s in splits:
        f = _f(s.get("factor", 1))
        tipo = "Agrupamento" if f < 1 else ("Desdobramento" if f > 1 else "—")
        rows.append([s.get("date", ""), s.get("ticker", ""), f,
                     s.get("ratio_str", ""), tipo])
    rows.sort(key=lambda r: (r[0], r[1]))
    _write_table(ws, headers, rows, fmts)
    _autosize(ws)


def sheet_aliases(wb, aliases):
    ws = wb.create_sheet("Aliases")
    headers = ["Data", "De", "Para", "Razão"]
    fmts = {"Razão": '0.000000'}
    rows = []
    for a in aliases:
        rows.append([a.get("date", ""), a.get("from_ticker", ""),
                     a.get("to_ticker", ""), _f(a.get("ratio", 1))])
    rows.sort(key=lambda r: r[0])
    _write_table(ws, headers, rows, fmts)
    _autosize(ws)


def sheet_precos(wb, prices):
    """Wide matrix: rows = month_end, cols = ticker, values = close."""
    ws = wb.create_sheet("Precos")
    by_me: dict[str, dict[str, float]] = defaultdict(dict)
    tickers: set[str] = set()
    for r in prices:
        by_me[r["month_end"]][r["ticker"]] = _f(r["close"])
        tickers.add(r["ticker"])
    months = sorted(by_me)
    tickers_sorted = sorted(tickers)
    headers = ["Mês fim"] + tickers_sorted
    fmts = {tk: NUM_FMTS["money_brl"] for tk in tickers_sorted}
    rows = []
    for me in months:
        row = [me] + [by_me[me].get(tk, "") for tk in tickers_sorted]
        rows.append(row)
    _write_table(ws, headers, rows, fmts)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.column_dimensions["A"].width = 12


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", default="investimentos_consolidado.xlsx")
    ap.add_argument("--twr", default="twr_monthly.csv")
    ap.add_argument("--full", default="twr_full.csv")
    ap.add_argument("--holdings-final", default="holdings_final.csv")
    ap.add_argument("--holdings-monthly", default="holdings_monthly.csv")
    ap.add_argument("--trades", default="trades.csv")
    ap.add_argument("--proventos", default="proventos.csv")
    ap.add_argument("--splits", default="splits.csv")
    ap.add_argument("--aliases", default="ticker_aliases.csv")
    ap.add_argument("--prices", default="prices_cache.csv")
    ap.add_argument("--rf-trades", default="rf_trades.csv")
    ap.add_argument("--rf-final", default="rf_final.csv")
    ap.add_argument("--rf-monthly", default="rf_holdings_monthly.csv")
    ap.add_argument("--us-trades", default="us_trades.csv")
    ap.add_argument("--us-final", default="us_final.csv")
    ap.add_argument("--us-summary", default="us_summary_monthly.csv")
    ap.add_argument("--us-divs", default="us_dividends.csv")
    ap.add_argument("--us-wh", default="us_withholding.csv")
    args = ap.parse_args()

    twr = _read_csv(args.twr)
    full = _read_csv(args.full)
    holdings_final = _read_csv(args.holdings_final)
    holdings_monthly = _read_csv(args.holdings_monthly)
    trades = _read_csv(args.trades)
    proventos = _read_csv(args.proventos)
    splits = _read_csv(args.splits)
    aliases = _read_csv(args.aliases)
    prices = _read_csv(args.prices)
    rf_trades = _read_csv(args.rf_trades)
    rf_final = _read_csv(args.rf_final)
    rf_monthly = _read_csv(args.rf_monthly)
    us_trades = _read_csv(args.us_trades)
    us_final = _read_csv(args.us_final)
    us_summary = _read_csv(args.us_summary)
    us_divs = _read_csv(args.us_divs)
    us_wh = _read_csv(args.us_wh)

    print(f"Loaded:")
    print(f"  twr_monthly:        {len(twr)} months")
    print(f"  twr_full:           {len(full)} months")
    print(f"  holdings_final:     {len(holdings_final)} rows")
    print(f"  holdings_monthly:   {len(holdings_monthly)} rows")
    print(f"  trades:             {len(trades)} rows")
    print(f"  proventos:          {len(proventos)} rows")
    print(f"  splits:             {len(splits)} rows")
    print(f"  aliases:            {len(aliases)} rows")
    print(f"  prices:             {len(prices)} rows")
    print(f"  rf_trades:          {len(rf_trades)} rows")
    print(f"  rf_final:           {len(rf_final)} rows")
    print(f"  rf_monthly:         {len(rf_monthly)} rows")

    wb = openpyxl.Workbook()
    # remove default
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    sheet_resumo(wb, twr, full, holdings_final, proventos, trades, rf_final, us_final)
    sheet_twr_mensal(wb, twr)
    sheet_vs_benchmarks(wb, full)
    sheet_holdings_final(wb, holdings_final)
    sheet_holdings_mes(wb, holdings_monthly)
    sheet_renda_fixa_atual(wb, rf_final)
    sheet_renda_fixa_historico(wb, rf_monthly)
    sheet_rf_trades(wb, rf_trades)
    sheet_us_atual(wb, us_final)
    sheet_us_historico(wb, us_summary)
    sheet_us_trades(wb, us_trades)
    sheet_us_dividends(wb, us_divs, us_wh)
    sheet_trades(wb, trades)
    sheet_proventos(wb, proventos)
    sheet_splits(wb, splits)
    sheet_aliases(wb, aliases)
    sheet_precos(wb, prices)

    wb.save(args.out)
    print(f"\nWrote {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
